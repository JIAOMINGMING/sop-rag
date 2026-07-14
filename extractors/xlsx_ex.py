"""Excel 解析器：每个 sheet 按行分块，表头行随每个 chunk 一起携带。

定位符：Sheet「名称」行N-M（从 1 开始，和用户在 Excel 里看到的一致）。
"""
from openpyxl import load_workbook

from . import make_chunk

ROWS_PER_CHUNK = 20


def row_text(cells):
    return " | ".join("" if c is None else str(c).strip() for c in cells).rstrip(" |")


def extract(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    stem = path.stem
    doc_title = stem.split("_", 1)[1] if "_" in stem else stem
    chunks, warnings = [], []
    for ws in wb.worksheets:
        rows = [(i, row) for i, row in enumerate(ws.iter_rows(values_only=True), start=1)
                if any(c is not None and str(c).strip() for c in row)]
        if not rows:
            continue
        header_idx, header = rows[0]
        header_line = row_text(header)
        data = rows[1:]
        if not data:                              # 只有表头的 sheet：整体作一个 chunk
            chunks.append(make_chunk(path, doc_title,
                                     f"Sheet「{ws.title}」行{header_idx}",
                                     f"Sheet {ws.title}", header_line))
            continue
        for i in range(0, len(data), ROWS_PER_CHUNK):
            block = data[i:i + ROWS_PER_CHUNK]
            first, last = block[0][0], block[-1][0]
            body = "\n".join(row_text(r) for _, r in block)
            text = f"表头: {header_line}\n{body}"
            chunks.append(make_chunk(path, doc_title,
                                     f"Sheet「{ws.title}」行{first}-{last}",
                                     f"Sheet {ws.title}", text))
    wb.close()
    if not chunks:
        warnings.append("所有 sheet 均为空")
    return {"chunks": chunks, "mode": "sheet-rows", "warnings": warnings}
